"""
excel_filler.py
Single consumer solar quotation filler.
Months pre-filled in template — only fills units, consumer details, bill amount.
"""

import shutil
import os
from openpyxl import load_workbook
from openpyxl.styles import Alignment


def fill_excel(data: dict, template_path: str, output_path: str) -> str:

    if not os.path.exists(template_path):
        raise FileNotFoundError(f"Template not found: {template_path}")

    shutil.copy(template_path, output_path)

    try:
        wb = load_workbook(output_path)
        ws = wb.active
    except Exception as e:
        raise ValueError(f"Could not open template: {str(e)}")

    # ── Consumer Details ───────────────────────────────────────
    ws["D6"] = data.get("consumer_name", "")

    # Consumer No as text — prevents scientific notation
    consumer_no = str(data.get("consumer_no", "")).strip()
    cell = ws["D7"]
    cell.value = consumer_no
    cell.number_format = "@"

    ws["D8"] = float(data.get("fixed_charges") or 130)
    ws["D9"] = data.get("sanctioned_load", "")
    ws["D10"] = data.get("connection_type", "")
    # D11 stays 600 (default solar panel watt)

    # ── Monthly Units ──────────────────────────────────────────
    monthly = data.get("monthly_units", [])
    if not monthly:
        raise ValueError("No monthly unit data found.")

    fixed_charges = float(data.get("fixed_charges") or 130)
    bill_amount = float(data.get("bill_amount") or 0)
    units_list = [float(m.get("units", 0)) for m in monthly]

    # Rows 15–26 = 12 months (months already pre-filled in column C)
    start_row = 15
    for i, entry in enumerate(monthly[:12]):
        row = start_row + i
        try:
            units = float(entry.get("units", 0))
            # Keep zero if zero — don't skip
            ws.cell(row=row, column=4, value=units)  # D = units
        except Exception as e:
            print(f"Warning: could not fill row {row}: {str(e)}")
            continue

    # ── Bill Amount + Unit Cost in last month row ──────────────
    if bill_amount and monthly:
        last_row = start_row + len(monthly) - 1
        last_units = units_list[-1] if units_list[-1] > 0 else 1
        unit_cost = round((bill_amount - fixed_charges) / last_units, 2)
        ws.cell(row=last_row, column=5, value=float(bill_amount))  # E = bill amount
        ws.cell(row=last_row, column=6, value=float(unit_cost))    # F = unit cost

    try:
        wb.save(output_path)
    except Exception as e:
        raise IOError(f"Could not save file: {str(e)}")

    return output_path


def get_summary_from_excel(data: dict) -> dict:
    """
    Calculate solar summary using same formulas as Excel template.
    """
    try:
        monthly = data.get("monthly_units", [])
        if not monthly:
            return {}

        units_list = [float(m.get("units", 0)) for m in monthly]
        avg_units = sum(units_list) / len(units_list)

        fixed_charges = float(data.get("fixed_charges") or 130)
        bill_amount = float(data.get("bill_amount") or 0)

        # Same as Excel: =(avg*12*1.1)/1400
        kw_required = (avg_units * 12 * 1.1) / 1400

        # Same as Excel: =kw/600*1000
        solar_panels_raw = kw_required / 600 * 1000

        # Same as Excel: =ROUND(panels,0)*600/1000
        solar_capacity = round(solar_panels_raw, 0) * 600 / 1000

        # Number of panels
        number_of_panels = int(round(solar_capacity / 600 * 1000, 0))

        # Unit cost from latest bill
        last_units = units_list[-1] if units_list[-1] > 0 else avg_units
        unit_cost = 0
        if bill_amount and last_units > 0:
            unit_cost = round((bill_amount - fixed_charges) / last_units, 2)

        # Yearly savings (solar covers ~90% of units)
        yearly_savings = round(avg_units * 12 * unit_cost * 0.9, 0)

        # Installation cost (₹45,000 per kWp)
        installation_cost = solar_capacity * 45000

        # Payback period in years
        payback_years = round(installation_cost / yearly_savings, 1) if yearly_savings > 0 else 0

        return {
            "avg_units":         round(avg_units, 2),
            "kw_required":       round(kw_required, 3),
            "solar_capacity":    solar_capacity,
            "panels_required":   number_of_panels,
            "bill_amount":       bill_amount,
            "unit_cost":         unit_cost,
            "yearly_savings":    int(yearly_savings),
            "installation_cost": int(installation_cost),
            "payback_years":     payback_years,
        }

    except Exception as e:
        print(f"Warning: summary calculation failed: {str(e)}")
        return {}